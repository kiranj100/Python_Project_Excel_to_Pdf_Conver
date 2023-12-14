import pandas as pd
# this is for show all files in folders to set path
import glob
import openpyxl
from fpdf import FPDF
# this is special for file path library
from pathlib import Path

file_path = glob.glob("Invoices/*.xlsx")
# print(file_path)

for second_file in file_path:

    pdf = FPDF(orientation="P",unit="mm",format="A4")
    pdf.add_page()

    # stem give us this 10001-2023.1.18 output and cut the .extension like .xlsx
    filename = Path(second_file).stem

        # it's give the list like['10001,'2023.1.18'] and zero position is 10001
    #invoice_No = filename.split("-")[0]

    # This is normal way
    # Date = filename.split("-")[1]

    # python distribute value accordingly to variables
    invoice_no, Date = filename.split("-")

    pdf.set_font(family="Times", style="B", size=18)
    pdf.cell(w=0, h=8, txt=f"Invoice No.{invoice_no}", ln=1)

    pdf.set_font(family="Times", style="B", size=18)
    pdf.cell(w=0, h=8, txt=f"Date: {Date}",ln=1)

    # pandas used read_excel to read .excel files
    df = pd.read_excel(second_file, sheet_name="Sheet 1")

    #add a header
    column = df.columns
    column = [item.replace("_"," ").title() for item in column]

    pdf.set_font(family="Times", style="B", size=10)
    pdf.set_text_color(100, 50, 100)
    pdf.cell(w=30, h=8, txt=str(column[0]), border=1)
    pdf.cell(w=70, h=8, txt=str(column[1]), border=1)
    pdf.cell(w=40, h=8, txt=str(column[2]), border=1)
    pdf.cell(w=30, h=8, txt=str(column[3]), border=1)
    pdf.cell(w=25, h=8, txt=str(column[4]), border=1, ln=1)

    #add rows to table
    for index,row in df.iterrows():
        pdf.set_font(family="Times",size=10)
        pdf.set_text_color(100,50,100)
        pdf.cell(w=30, h=8, txt=str(row["product_id"]),border=1)
        pdf.cell(w=70, h=8, txt=str(row["product_name"]),border=1)
        pdf.cell(w=40, h=8, txt=str(row["amount_purchased"]),border=1)
        pdf.cell(w=30, h=8, txt=str(row["price_per_unit"]),border=1)
        pdf.cell(w=25, h=8, txt=str(row["total_price"]),border=1,ln=1)
    total_sum = df["total_price"].sum()
    pdf.set_font(family="Times", size=10)
    pdf.set_text_color(100, 50, 100)
    pdf.cell(w=30, h=8, txt="", border=1)
    pdf.cell(w=70, h=8, txt="", border=1)
    pdf.cell(w=40, h=8, txt="", border=1)
    pdf.cell(w=30, h=8, txt="", border=1)
    pdf.cell(w=25, h=8, txt=str(total_sum), border=1, ln=1)

    # Add total sum in sentence
    pdf.set_font(family="Times", size=15,style="B")
    pdf.cell(w=70, h=8, txt=f"The Total Price is {total_sum}", ln=1)

    # Add company name and logo
    pdf.set_font(family="Times", style="B", size=18)
    pdf.cell(w=32, h=8,txt=f"PythonHow")
    pdf.image("pythonhow.png", w=10)

    # files name is this 10001-2023.1.18
    pdf.output(f"PDFS/{filename}.pdf")


